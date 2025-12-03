from openpyxl import load_workbook
import csv
import os

def xlsx_to_csv(xlsx_path):
    wb = load_workbook(xlsx_path)
    ws = wb.active

    csv_path = os.path.splitext(xlsx_path)[0] + ".csv"

    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(row)

    print("Saved:", csv_path)


xlsx_to_csv("elemnts_from_poetries.xlsx")
